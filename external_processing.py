import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl
import statistics

raw_data = xlrd.open_workbook("raw_data.xlsx")
sheet_rd = raw_data.sheet_by_index(0)

data_pipeline = openpyxl.load_workbook(filename='data_pipeline.xlsx')
sheet_dp = data_pipeline['Sheet1']


# sheep_position
sheep_longitude = sheet_rd.cell_value(10, 0)
sheep_latitude = sheet_rd.cell_value(10, 1)
sheep_position = "Longitude: " + str(sheep_longitude) + " Latitude: " + str(sheep_latitude)
sheet_dp.cell(row=2, column=1, value=sheep_position)
data_pipeline.save('data_pipeline.xlsx')

# sheep_velocity
sheep_velocity = sheet_rd.cell_value(10, 2)
sheet_dp.cell(row=2, column=2, value=sheep_velocity)
data_pipeline.save('data_pipeline.xlsx')

# drone_position
drone_longitude = sheet_rd.cell_value(10, 5)
drone_latitude = sheet_rd.cell_value(10,6)
drone_position = "Longitude: " + str(drone_longitude) + " Latitude: " + str(drone_latitude)
sheet_dp.cell(row=2, column=3, value=drone_position)
data_pipeline.save('data_pipeline.xlsx')

# drone_velocity
drone_velocity = sheet_rd.cell_value(10, 7)
sheet_dp.cell(row=2, column=4, value=drone_velocity)
data_pipeline.save('data_pipeline.xlsx')

# drone_battery
drone_battery = sheet_rd.cell_value(10, 9)
sheet_dp.cell(row=2, column=5, value=drone_battery)
data_pipeline.save('data_pipeline.xlsx')

# drone_actions
drone_actions = sheet_rd.cell_value(10, 10)
sheet_dp.cell(row=2, column=6, value=drone_actions)
data_pipeline.save('data_pipeline.xlsx')

# sheep_psych
heart_rate = []
breathing_rate = []
for i in range(10):
    heart_rate.append(sheet_rd.cell_value(i+1, 3))
for i in range(10):
    breathing_rate.append(sheet_rd.cell_value(i+1, 4))
avg_hr = statistics.mean(heart_rate)
avg_br = statistics.mean(breathing_rate)
if avg_hr > 100 or avg_br > 40:
    sheep_psych = "stressed"
else:
    sheep_psych = "not stressed"
sheet_dp.cell(row=2, column=7, value=sheep_psych)
data_pipeline.save('data_pipeline.xlsx')

# drone_sheep_dynamic
if sheep_velocity == 0 and avg_hr < 80 and avg_br < 34:
    drone_sheep_dynamic = "unnoticed"
elif sheep_velocity == 0 and avg_hr > 80 or avg_br > 34:
    drone_sheep_dynamic = "alert"
elif sheep_velocity > 0 and avg_hr < 100 and avg_br < 40:
    drone_sheep_dynamic = "herded"
elif sheep_velocity > 0 and avg_hr > 100 or avg_br > 40:
    drone_sheep_dynamic = "fear"
else:
    drone_sheep_dynamic = "unknown"
sheet_dp.cell(row=2, column=8, value=drone_sheep_dynamic)
data_pipeline.save('data_pipeline.xlsx')

# sheep_sheep_dynamic
sheep_sheep_dynamic = sheet_rd.cell_value(10, 11)
sheet_dp.cell(row=2, column=9, value=sheep_sheep_dynamic)
data_pipeline.save('data_pipeline.xlsx')

# human_drone_dynamic
sheet_dp.cell(row=2, column=10, value=drone_actions)
data_pipeline.save('data_pipeline.xlsx')

# drone_altitude
drone_altitude = sheet_rd.cell_value(10, 8)
sheet_dp.cell(row=2, column=11, value=drone_altitude)
data_pipeline.save('data_pipeline.xlsx')